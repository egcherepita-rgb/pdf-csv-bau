import io
import os
import re
import time
import json
import threading
from uuid import uuid4
from collections import OrderedDict
from typing import List, Tuple, Dict, Optional, Any

import fitz  # PyMuPDF
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import Response, HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles

try:
    import openpyxl
except Exception:
    openpyxl = None


app = FastAPI(
    title="Бауцентр • PDF → XLSX (АРТИКУЛ / ШТ / ПЛОЩАДЬ)",
    version="1.0.4",
)

# Static files (logo etc.)
app.mount("/static", StaticFiles(directory="static"), name="static")

# -------------------------
# Regex / helpers
# -------------------------
RX_SIZE = re.compile(r"\b\d{2,}[xх×]\d{2,}(?:[xх×]\d{1,})?\b", re.IGNORECASE)
RX_MM = re.compile(r"мм", re.IGNORECASE)
RX_WEIGHT = re.compile(r"\b\d+(?:[.,]\d+)?\s*кг\.?\b", re.IGNORECASE)

RX_MONEY_LINE = re.compile(r"^\d+(?:[ \u00a0]\d{3})*(?:[.,]\d+)?\s*₽$")
RX_INT = re.compile(r"^\d+$")
RX_ANY_RUB = re.compile(r"₽")

RX_PRICE_QTY_SUM = re.compile(
    r"(?P<price>\d+(?:[ \u00a0]\d{3})*(?:[.,]\d+)?)\s*₽\s+"
    r"(?P<qty>\d{1,4})\s+"
    r"(?P<sum>\d+(?:[ \u00a0]\d{3})*(?:[.,]\d+)?)\s*₽",
    re.IGNORECASE,
)

RX_DIMS_ANYWHERE = re.compile(
    r"\s*\d{1,4}[xх×]\d{1,4}(?:[xх×]\d{1,5})?\s*мм\.?\s*",
    re.IGNORECASE,
)

# Площадь: "1.23 м2", "1,23 м²", "1.23 кв. м"
RX_AREA = re.compile(
    r"(?P<val>\d+(?:[.,]\d+)?)\s*(?:м2|м²|кв\.?\s*м|квм)\b",
    re.IGNORECASE,
)

RX_FLOAT_ONLY = re.compile(r"^\d+(?:[.,]\d+)?$")


def normalize_space(s: str) -> str:
    s = (s or "").replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalize_key(name: str) -> str:
    s = normalize_space(name).lower()
    s = s.replace("×", "x").replace("х", "x")
    s = RX_DIMS_ANYWHERE.sub(" ", s)
    s = normalize_space(s)
    return s


def strip_dims_anywhere(name: str) -> str:
    name = normalize_space(name)
    name2 = RX_DIMS_ANYWHERE.sub(" ", name)
    return normalize_space(name2)


# -------------------------
# Артикулы (Art1.xlsx) — кастомные для Бауцентра
# ENV:
#   ART_XLSX_PATH=/path/Art1.xlsx
#   ART_VALUE_COLUMN=BAU  (или "Артикул")
# -------------------------
def load_article_map() -> Tuple[Dict[str, str], str]:
    """Загружает соответствие 'Товар' -> 'Артикул' из Art1.xlsx.

    Важно: в файле часто встречается Артикул = 0 — это валидное значение, его НЕ пропускаем.
    Путь можно задать через ENV ART_XLSX_PATH. Если указан относительный путь, пробуем
    также рядом с main.py и в текущей папке запуска.
    """
    if openpyxl is None:
        return {}, "openpyxl_not_installed"

    env_path = os.getenv("ART_XLSX_PATH", "Art1.xlsx")
    art_value_col_name = normalize_space(os.getenv("ART_VALUE_COLUMN", "Артикул"))

    candidates = []
    if env_path:
        candidates.append(env_path)
        if not os.path.isabs(env_path):
            # рядом с файлом приложения
            try:
                here = os.path.dirname(os.path.abspath(__file__))
                candidates.append(os.path.join(here, env_path))
            except Exception:
                pass
            # текущая директория
            candidates.append(os.path.join(os.getcwd(), env_path))

    path = next((p for p in candidates if p and os.path.exists(p)), None)
    if not path:
        return {}, f"file_not_found:{env_path}"

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        return {}, f"cannot_open:{e}"

    header = [normalize_space(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]

    товар_col = None
    art_col = None

    for idx, h in enumerate(header, start=1):
        if h.lower() == "товар":
            товар_col = idx
        if art_value_col_name and h.lower() == art_value_col_name.lower():
            art_col = idx

    if товар_col is None:
        товар_col = 1

    if art_col is None:
        for idx, h in enumerate(header, start=1):
            if h.lower() == "артикул":
                art_col = idx
                break

    if art_col is None:
        art_col = 2 if ws.max_column >= 2 else 1

    def art_to_str(val: Any) -> str:
        if val is None:
            return ""
        # числа из Excel могут быть float
        if isinstance(val, (int,)):
            return str(val)
        if isinstance(val, float):
            if abs(val - int(val)) < 1e-9:
                return str(int(val))
            return str(val).replace(",", ".")
        s = normalize_space(str(val))
        # если это "670000078.0" → "670000078"
        if re.fullmatch(r"\d+\.0", s):
            return s[:-2]
        return s

    m: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        товар = ws.cell(r, товар_col).value
        арт = ws.cell(r, art_col).value

        if товар is None:
            continue

        товар_s = normalize_space(str(товар))
        арт_s = art_to_str(арт)

        # Артикул может быть "0" — это валидно. Пропускаем только если совсем пусто.
        if not товар_s or арт_s == "":
            continue

        m[normalize_key(товар_s)] = арт_s

    return m, "ok"


ARTICLE_MAP, ARTICLE_MAP_STATUS = load_article_map()

# -------------------------
# Счетчик конвертаций
# -------------------------
from threading import Lock

COUNTER_FILE = os.getenv("COUNTER_FILE", "conversions.count")
_counter_lock = Lock()


def _read_counter() -> int:
    try:
        with open(COUNTER_FILE, "r", encoding="utf-8") as f:
            return int((f.read() or "0").strip())
    except Exception:
        return 0


def _write_counter(v: int) -> None:
    tmp = COUNTER_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        f.write(str(v))
    os.replace(tmp, COUNTER_FILE)


def increment_counter() -> int:
    with _counter_lock:
        v = _read_counter() + 1
        _write_counter(v)
        return v


def get_counter() -> int:
    return _read_counter()


# -------------------------
# Async jobs (progress) — FILE storage
# -------------------------
JOB_DIR = os.getenv("JOB_DIR", "/tmp/pdf2xlsx_bau_jobs")
JOB_TTL_SEC = int(os.getenv("JOB_TTL_SEC", str(30 * 60)))  # seconds


def _ensure_job_dir() -> None:
    os.makedirs(JOB_DIR, exist_ok=True)


def _job_json_path(job_id: str) -> str:
    return os.path.join(JOB_DIR, f"{job_id}.json")


def _job_result_path(job_id: str) -> str:
    return os.path.join(JOB_DIR, f"{job_id}.xlsx")


def _write_json_atomic(path: str, data: dict) -> None:
    _ensure_job_dir()
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    os.replace(tmp, path)


def _read_json(path: str) -> Optional[dict]:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _cleanup_jobs() -> None:
    _ensure_job_dir()
    now = time.time()
    try:
        for name in os.listdir(JOB_DIR):
            if not (name.endswith(".json") or name.endswith(".xlsx")):
                continue
            p = os.path.join(JOB_DIR, name)
            try:
                if now - os.path.getmtime(p) > JOB_TTL_SEC:
                    os.remove(p)
            except Exception:
                pass
    except Exception:
        pass


def _set_job(job_id: str, **kwargs) -> None:
    _cleanup_jobs()
    path = _job_json_path(job_id)
    data = _read_json(path) or {}
    data.update(kwargs)
    _write_json_atomic(path, data)


def _get_job(job_id: str) -> Optional[dict]:
    _cleanup_jobs()
    return _read_json(_job_json_path(job_id))


def _set_job_result(job_id: str, xlsx_bytes: bytes) -> None:
    _cleanup_jobs()
    _ensure_job_dir()
    p = _job_result_path(job_id)
    tmp = p + ".tmp"
    with open(tmp, "wb") as f:
        f.write(xlsx_bytes)
    os.replace(tmp, p)


def _get_job_result(job_id: str) -> Optional[bytes]:
    p = _job_result_path(job_id)
    try:
        with open(p, "rb") as f:
            return f.read()
    except Exception:
        return None


# -------------------------
# Instruction media (optional)
# -------------------------
INSTRUCTION_VIDEO_PATH = os.getenv("INSTRUCTION_VIDEO_PATH", "instruction.mp4")

# -------------------------
# PDF parsing helpers
# -------------------------
def is_noise(line: str) -> bool:
    low = (line or "").strip().lower()
    if not low:
        return True
    if low.startswith("страница:"):
        return True
    if low.startswith("ваш проект"):
        return True
    if "проект создан" in low:
        return True
    if "развертка стены" in low:
        return True
    if "стоимость проекта" in low:
        return True
    # Заголовок таблицы из PDF: "ID Фото Товар Габариты Вес Цена за шт Кол-во Сумма"
    if low.startswith("id ") and "фото" in low and "товар" in low and "сумма" in low:
        return True
    return False


def is_totals_block(line: str) -> bool:
    low = (line or "").strip().lower()
    return (
        low.startswith("общий вес")
        or low.startswith("максимальный габарит заказа")
        or low.startswith("адрес:")
        or low.startswith("телефон:")
        or low.startswith("email")
    )


def money_to_number(line: str) -> int:
    s = normalize_space(line).replace("₽", "").strip()
    s = s.replace("\u00a0", " ")
    s = s.replace(" ", "")
    s = s.replace(",", ".")
    try:
        return int(float(s))
    except Exception:
        return -1


def is_project_total_only(line: str, prev_line: str = "") -> bool:
    if not RX_MONEY_LINE.fullmatch(normalize_space(line)):
        return False
    prev = normalize_space(prev_line).lower()
    if "стоимость проекта" in prev:
        return True
    v = money_to_number(line)
    return v >= 10000


def is_header_token(line: str) -> bool:
    low = normalize_space(line).lower().replace("–", "-").replace("—", "-")
    return low in {"id", "фото", "товар", "габариты", "вес", "цена за шт", "кол-во", "сумма", "площадь"}


def looks_like_dim_or_weight(line: str) -> bool:
    if RX_WEIGHT.search(line):
        return True
    if RX_SIZE.search(line) and RX_MM.search(line):
        return True
    return False


def looks_like_money_or_qty(line: str) -> bool:
    if RX_MONEY_LINE.fullmatch(line):
        return True
    if RX_INT.fullmatch(line):
        return True
    return False


def clean_name_from_buffer(buf: List[str]) -> str:
    filtered: List[str] = []
    for ln in buf:
        if is_noise(ln) or is_header_token(ln) or is_totals_block(ln):
            continue
        filtered.append(ln)

    # Убираем строки, которые содержат только числовой ID (встречается как отдельная строка "0")
    filtered = [ln for ln in filtered if not RX_INT.fullmatch(ln)]

    # Если первая строка начинается с ID + пробел + текст, убираем этот ID:
    # "670000078 Рельс несущий" -> "Рельс несущий"
    if filtered:
        filtered[0] = re.sub(r"^\d{3,}\s+", "", filtered[0]).strip()

    while filtered and (looks_like_dim_or_weight(filtered[-1]) or looks_like_money_or_qty(filtered[-1])):
        filtered.pop()

    name = normalize_space(" ".join(filtered))
    name = re.sub(r"^Фото\s*", "", name, flags=re.IGNORECASE).strip()
    name = re.sub(r"^Товар\s*", "", name, flags=re.IGNORECASE).strip()
    name = strip_dims_anywhere(name)
    return name


def _to_float(s: str) -> float:
    s = normalize_space(s).replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def extract_area_from_context(lines: List[str]) -> float:
    # 1) явные единицы
    for ln in lines:
        m = RX_AREA.search(ln)
        if m:
            return _to_float(m.group("val"))

    # 2) по слову "площадь"
    for idx, ln in enumerate(lines):
        low = ln.lower()
        if "площад" in low:
            mm = re.search(r"(\d+(?:[.,]\d+)?)", ln)
            if mm:
                return _to_float(mm.group(1))
            if idx + 1 < len(lines) and RX_FLOAT_ONLY.fullmatch(lines[idx + 1]):
                return _to_float(lines[idx + 1])

    return 0.0


# -------------------------
# Main parser: name -> (qty_sum, area_sum)
# -------------------------

def parse_items(pdf_bytes: bytes) -> Tuple[List[Tuple[str, int, float]], Dict[str, Any]]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    total_pages = doc.page_count

    ordered: "OrderedDict[str, Dict[str, Any]]" = OrderedDict()

    stats = {
        "pages": 0,
        "total_pages": total_pages,
        "processed_pages": 0,
        "items_found": 0,
        "anchors_inline": 0,
        "anchors_multiline": 0,
        "article_map_size": len(ARTICLE_MAP),
        "article_map_status": ARTICLE_MAP_STATUS,
        "parser": "id_segment_v3",
    }

    def flush_segment(seg_lines: List[str]) -> None:
        if not seg_lines:
            return

        raw = [normalize_space(x) for x in seg_lines if normalize_space(x)]
        if not raw:
            return

        # Первая строка сегмента почти всегда содержит ID
        if not RX_INT.fullmatch(raw[0]):
            return

        work = raw[1:]

        # Отрезаем служебный хвост, если он внезапно попал в сегмент
        cut_idx = None
        for idx, ln in enumerate(work):
            if is_totals_block(ln) or is_noise(ln):
                cut_idx = idx
                break
        if cut_idx is not None:
            work = work[:cut_idx]

        if not work:
            return

        qty = 0
        anchor_kind = None

        joined = " ".join(work)
        m_inline = RX_PRICE_QTY_SUM.search(joined)
        if m_inline:
            try:
                qty = int(m_inline.group("qty"))
            except Exception:
                qty = 0
            anchor_kind = "inline"

        if not (1 <= qty <= 500):
            for i, ln in enumerate(work):
                if RX_MONEY_LINE.fullmatch(ln):
                    if i + 2 < len(work) and RX_INT.fullmatch(work[i + 1]) and RX_MONEY_LINE.fullmatch(work[i + 2]):
                        try:
                            qty = int(work[i + 1])
                        except Exception:
                            qty = 0
                        anchor_kind = "multiline"
                        break

        if not (1 <= qty <= 500):
            return

        # Название = строки до первой строки с габаритами/весом/ценой
        name_lines: List[str] = []
        for ln in work:
            if looks_like_dim_or_weight(ln) or RX_MONEY_LINE.fullmatch(ln):
                break
            if RX_INT.fullmatch(ln):
                continue
            if is_header_token(ln) or is_noise(ln) or is_totals_block(ln):
                continue
            name_lines.append(ln)

        name = strip_dims_anywhere(normalize_space(" ".join(name_lines)))
        name = re.sub(r"^Фото\s*", "", name, flags=re.IGNORECASE).strip()
        name = re.sub(r"^Товар\s*", "", name, flags=re.IGNORECASE).strip()

        if not name:
            return

        area = extract_area_from_context(work)

        if name not in ordered:
            ordered[name] = {"qty": 0, "area": 0.0}
        ordered[name]["qty"] += qty
        ordered[name]["area"] += float(area or 0.0)
        stats["items_found"] += 1
        if anchor_kind == "inline":
            stats["anchors_inline"] += 1
        else:
            stats["anchors_multiline"] += 1

    for page in doc:
        stats["pages"] += 1
        stats["processed_pages"] += 1

        txt = page.get_text("text") or ""
        if "₽" not in txt and "ID" not in txt and not re.search(r"\b\d{6,}\b", txt):
            continue

        lines = [normalize_space(x) for x in txt.splitlines()]
        lines = [x for x in lines if x]
        if not lines:
            continue

        segment: List[str] = []
        in_table = False

        for idx, line in enumerate(lines):
            prev = lines[idx - 1] if idx > 0 else ""

            if is_noise(line) or is_header_token(line):
                # Заголовок таблицы может быть без отдельной строки с "ID Фото..."
                if "id" in line.lower() and "товар" in line.lower():
                    in_table = True
                continue

            # Начало новой позиции по ID
            if RX_INT.fullmatch(line) and (len(line) >= 6 or line == '0'):
                in_table = True
                if segment:
                    flush_segment(segment)
                segment = [line]
                continue

            if not in_table:
                continue

            # После блока итогов позиции заканчиваются. Саму сумму проекта здесь не ловим,
            # иначе можно ошибочно оборвать строку с большой суммой по позиции.
            if is_totals_block(line):
                if segment:
                    flush_segment(segment)
                    segment = []
                in_table = False
                continue

            if segment:
                segment.append(line)

        if segment:
            flush_segment(segment)

    out_rows: List[Tuple[str, int, float]] = []
    for name, v in ordered.items():
        out_rows.append((name, int(v.get("qty") or 0), float(v.get("area") or 0.0)))

    return out_rows, stats


# -------------------------
# XLSX output (АРТИКУЛ / ШТ / ПЛОЩАДЬ)
# - если артикула нет в Art1.xlsx → пишем наименование товара
# - если площадь = 0 → пустая ячейка
# -------------------------
def make_xlsx(rows: List[Tuple[str, int, float]]) -> bytes:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BAU"

    ws.append(["АРТИКУЛ", "ШТ", "ПЛОЩАДЬ"])

    for name, qty, area in rows:
        art = ARTICLE_MAP.get(normalize_key(name), "")
        art_out = name if (not art or str(art).strip() == '0') else art  # fallback на наименование, если нет артикула или он = 0

        area_cell = float(area) if area and float(area) > 0 else None  # None => пусто в Excel

        ws.append([art_out, int(qty or 0), area_cell])

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14

    for cell in ws["B"][1:]:
        cell.number_format = "0"
    for cell in ws["C"][1:]:
        cell.number_format = "0.00"

    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# -------------------------
# UI (компактная версия)
# -------------------------
HOME_HTML = """<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Бауцентр • PDF → XLSX</title>
  <style>
    :root{
      --bg1:#0a1020;
      --bg2:#0b0f17;
      --stroke: rgba(255,255,255,.12);
      --text:#eef2ff;
      --muted: rgba(238,242,255,.72);
      --shadow: 0 30px 80px rgba(0,0,0,.45);
      --accent:#ffd33d;
      --ok:#5CFF9A;
      --err:#ff6b7a;
    }
    *{box-sizing:border-box;}
    html,body{height:100%; overflow:hidden;}
    body{
      margin:0;
      font-family: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
      color:var(--text);
      background:
        radial-gradient(1000px 600px at 15% 10%, #1a2c66 0%, transparent 60%),
        radial-gradient(900px 500px at 85% 15%, #4a2a1a 0%, transparent 55%),
        linear-gradient(180deg, var(--bg1), var(--bg2));
    }

    .topbar{
      position:fixed;
      inset: 20px 18px auto 18px;
      display:flex;
      align-items:center;
      justify-content:center;
      z-index:10;
      pointer-events:none;
    }
    .logo-shell{
      display:flex;
      align-items:center;
      justify-content:center;
      width:min(420px, 92vw);
      min-height:96px;
      padding:12px 22px;
      border:1px solid rgba(255,255,255,.08);
      border-radius:22px;
      background: linear-gradient(180deg, rgba(255,255,255,.06), rgba(255,255,255,.025));
      backdrop-filter: blur(10px);
      box-shadow: 0 18px 45px rgba(0,0,0,.25);
    }
    .logo{
      display:block;
      width:min(290px, 68vw);
      height:auto;
      filter: drop-shadow(0 10px 24px rgba(0,0,0,.35));
      opacity:.99;
    }

    .wrap{
      height:100vh;
      display:flex;
      align-items:center;
      justify-content:center;
      padding: 18px 20px 20px;
      padding-top: 132px;
    }
    .card{
      width:min(940px, 100%);
      border:1px solid var(--stroke);
      background: linear-gradient(180deg, rgba(255,255,255,.08), rgba(255,255,255,.04));
      border-radius: 24px;
      box-shadow: var(--shadow);
      padding: 22px;
      backdrop-filter: blur(10px);
    }
    .head{
      display:flex;
      gap:14px;
      align-items:flex-start;
      justify-content:space-between;
      flex-wrap:wrap;
      margin-bottom: 12px;
    }
    h1{
      margin:0;
      font-size:30px;
      line-height:1.08;
      letter-spacing:.2px;
    }
    .sub{
      margin-top:8px;
      color:var(--muted);
      font-size:14px;
      line-height:1.45;
      max-width:760px;
    }
    .pill{
      font-size:12px;
      color:var(--muted);
      border:1px solid var(--stroke);
      padding:6px 10px;
      border-radius:999px;
      white-space:nowrap;
      background: rgba(255,255,255,.04);
    }

    .drop{
      margin-top: 14px;
      border: 1px dashed rgba(255,255,255,.22);
      background: rgba(0,0,0,.12);
      border-radius: 20px;
      padding: 18px;
      display:flex;
      gap: 14px;
      align-items:center;
      justify-content:space-between;
      flex-wrap:wrap;
      transition: border-color .15s ease, background .15s ease;
    }
    .drop-left{
      display:flex;
      gap: 12px;
      align-items:center;
      flex: 1 1 420px;
      min-width: 0;
    }
    .icon{
      width:44px;height:44px;
      border-radius: 14px;
      background: rgba(255,211,61,.15);
      border: 1px solid rgba(255,211,61,.35);
      display:flex;
      align-items:center;
      justify-content:center;
      font-weight:900;
      color: var(--accent);
      user-select:none;
      flex: 0 0 44px;
    }
    .drop-title{
      font-weight:800;
      font-size:16px;
      line-height:1.25;
      word-break:break-word;
    }
    .drop-hint{
      color:var(--muted);
      font-size:13px;
      margin-top:4px;
    }

    input[type=file]{display:none;}
    .btn{
      pointer-events:auto;
      display:inline-flex;
      align-items:center;
      justify-content:center;
      gap:10px;
      padding: 10px 14px;
      min-height: 46px;
      border:0;
      border-radius: 14px;
      cursor:pointer;
      font-weight: 900;
      background: var(--accent);
      color:#1a1a1a;
      box-shadow: 0 10px 26px rgba(255,211,61,.18);
      transition: transform .08s ease;
      text-decoration:none;
    }
    .btn:active{transform: translateY(1px);}
    .btn.secondary{
      background: rgba(255,255,255,.08);
      color: var(--text);
      border: 1px solid var(--stroke);
      box-shadow:none;
    }
    .btn:disabled{opacity:.55; cursor:not-allowed;}

    .status{
      margin-top: 12px;
      font-size: 14px;
      color: var(--muted);
      text-align:center;
      white-space:pre-wrap;
      min-height: 20px;
    }
    .status.ok{color: var(--ok);}
    .status.err{color: var(--err);}

    .bar{
      margin-top: 10px;
      height: 10px;
      background: rgba(255,255,255,.08);
      border: 1px solid rgba(255,255,255,.10);
      border-radius: 999px;
      overflow:hidden;
    }
    .bar > div{
      height:100%;
      width:0%;
      background: rgba(255,211,61,.9);
      transition: width .25s ease;
    }

    .corner{
      position: fixed;
      right: 14px;
      bottom: 12px;
      font-size: 12px;
      color: var(--muted);
      opacity:.9;
      padding: 7px 10px;
      border-radius:999px;
      background: rgba(0,0,0,.16);
      border: 1px solid rgba(255,255,255,.08);
      backdrop-filter: blur(6px);
    }

    @media (max-width: 760px){
      html,body{overflow:auto;}
      .topbar{
        position:relative;
        inset:auto;
        padding: 14px 14px 0;
      }
      .wrap{
        height:auto;
        min-height: calc(100vh - 20px);
        padding: 14px 14px 22px;
      }
      .logo-shell{
        width:min(330px, 94vw);
        min-height:80px;
        padding:10px 16px;
      }
      .logo{
        width:min(230px, 72vw);
      }
      .card{
        padding:18px;
        border-radius:20px;
      }
      h1{
        font-size:24px;
      }
      .drop{
        padding:16px;
      }
      .drop-left{
        align-items:flex-start;
        flex-basis:100%;
      }
      .btn, .btn.secondary{
        width:100%;
      }
      .corner{
        position: static;
        width: fit-content;
        margin: 12px auto 0;
      }
    }
  </style>
</head>
<body>
  <div class="topbar">
    <div class="logo-shell">
      <img class="logo" src="/static/logo.png" alt="Бауцентр" />
    </div>
  </div>

  <div class="wrap">
    <div class="card">
      <div class="head">
        <div>
          <h1>Конвертация PDF → XLSX</h1>
          <div class="sub">На выходе Excel с 3 колонками: <b>АРТИКУЛ/НАИМЕНОВАНИЕ</b>, <b>ШТ</b>, <b>ПЛОЩАДЬ</b> (пустая, если нет).</div>
        </div>
        <div class="pill">bau.pdfcsv.ru</div>
      </div>

      <div id="drop" class="drop">
        <div class="drop-left">
          <div class="icon">PDF</div>
          <div>
            <div class="drop-title" id="fname">Перетащи PDF сюда или выбери файл</div>
            <div class="drop-hint">Поддерживается только *.pdf</div>
          </div>
        </div>

        <div style="display:flex; gap:10px; flex-wrap:wrap;">
          <label class="btn secondary" for="pdf">Выбрать PDF</label>
          <button id="btn" class="btn" disabled>Скачать XLSX</button>
        </div>

        <input id="pdf" type="file" accept="application/pdf,.pdf" />
      </div>

      <div class="bar" aria-hidden="true"><div id="bar"></div></div>
      <div id="status" class="status"></div>
    </div>
  </div>

  <div class="corner" id="counter">…</div>

<script>
  const input = document.getElementById('pdf');
  const btn = document.getElementById('btn');
  const statusEl = document.getElementById('status');
  const fnameEl = document.getElementById('fname');
  const barEl = document.getElementById('bar');
  const drop = document.getElementById('drop');

  function ok(msg){ statusEl.className='status ok'; statusEl.textContent=msg; }
  function err(msg){ statusEl.className='status err'; statusEl.textContent=msg; }
  function neutral(msg){ statusEl.className='status'; statusEl.textContent=msg||''; }
  function setBar(p){ barEl.style.width = Math.max(0, Math.min(100, p)) + '%'; }

  async function loadCounter(){
    try {
      const r = await fetch('/stats');
      if (!r.ok) return;
      const j = await r.json();
      if (typeof j.conversions === 'number') {
        document.getElementById('counter').textContent = String(j.conversions);
      }
    } catch(e) {}
  }
  loadCounter();

  function setFile(f){
    if (!f) {
      btn.disabled = true;
      fnameEl.textContent = 'Перетащи PDF сюда или выбери файл';
      setBar(0);
      neutral('');
      return;
    }
    btn.disabled = false;
    fnameEl.textContent = 'Выбран файл: ' + f.name;
    setBar(0);
    neutral('Готов к конвертации');
  }

  input.addEventListener('change', () => setFile(input.files && input.files[0]));

  ;['dragenter','dragover'].forEach(ev => drop.addEventListener(ev, e => {
    e.preventDefault(); e.stopPropagation();
    drop.style.borderColor = 'rgba(255,211,61,.55)';
    drop.style.background = 'rgba(0,0,0,.16)';
  }));
  ;['dragleave','drop'].forEach(ev => drop.addEventListener(ev, e => {
    e.preventDefault(); e.stopPropagation();
    drop.style.borderColor = 'rgba(255,255,255,.22)';
    drop.style.background = 'rgba(0,0,0,.12)';
  }));
  drop.addEventListener('drop', e => {
    const f = e.dataTransfer.files && e.dataTransfer.files[0];
    if (f) {
      input.files = e.dataTransfer.files;
      setFile(f);
    }
  });

  btn.addEventListener('click', async () => {
    const f = input.files && input.files[0];
    if (!f) return;

    btn.disabled = true;
    setBar(10);
    const start = Date.now();

    try {
      const fd = new FormData();
      fd.append('file', f);

      neutral('Загружаю PDF…');
      const r = await fetch('/extract_async', { method: 'POST', body: fd });
      if (!r.ok) throw new Error(await r.text());

      const data = await r.json();
      const job_id = data.job_id;

      while (true) {
        const s = await fetch('/job/' + job_id);
        if (!s.ok) throw new Error(await s.text());
        const j = await s.json();

        const sec = Math.floor((Date.now() - start) / 1000);
        let msg = (j.message || 'Обработка…') + ' • ' + sec + ' сек';
        if (j.total_pages && j.processed_pages) {
          msg += ' • страниц: ' + j.processed_pages + '/' + j.total_pages;
          setBar(10 + (j.processed_pages / j.total_pages) * 70);
        }

        neutral(msg);

        if (j.status === 'done') {
          setBar(92);
          const dl = await fetch('/job/' + job_id + '/download');
          if (!dl.ok) throw new Error(await dl.text());
          const blob = await dl.blob();

          const filename = (j.filename || ((f.name || 'items.pdf').replace(/\\.pdf$/i, '') + '.xlsx'));
          const url = URL.createObjectURL(blob);
          const a = document.createElement('a');
          a.href = url;
          a.download = filename;
          document.body.appendChild(a);
          a.click();
          a.remove();
          URL.revokeObjectURL(url);

          setBar(100);
          ok('Готово! XLSX скачан: ' + filename);
          loadCounter();
          break;
        }

        if (j.status === 'error') throw new Error(j.message || 'Ошибка обработки');
        await new Promise(res => setTimeout(res, 600));
      }

    } catch (e) {
      setBar(0);
      err('Ошибка: ' + String(e.message || e));
    } finally {
      btn.disabled = !(input.files && input.files[0]);
    }
  });
</script>
</body>
</html>
"""


# -------------------------
# Endpoints
# -------------------------
@app.get("/stats")
def stats():
    return {"conversions": get_counter()}


@app.get("/health")
def health():
    try:
        _ensure_job_dir()
        job_files = len([x for x in os.listdir(JOB_DIR) if x.endswith(".json")])
    except Exception:
        job_files = -1

    return {
        "status": "ok",
        "article_map_size": len(ARTICLE_MAP),
        "article_map_status": ARTICLE_MAP_STATUS,
        "art_value_column": os.getenv("ART_VALUE_COLUMN", ""),
        "art_xlsx_path": os.getenv("ART_XLSX_PATH", ""),
        "conversions": get_counter(),
        "job_dir": JOB_DIR,
        "job_files": job_files,
        "openpyxl": bool(openpyxl is not None),
    }


@app.api_route("/", methods=["GET", "HEAD"], response_class=HTMLResponse)
def home():
    return HOME_HTML


@app.api_route("/instruction.mp4", methods=["GET", "HEAD"])
def instruction_video():
    if not os.path.exists(INSTRUCTION_VIDEO_PATH):
        raise HTTPException(status_code=404, detail="instruction.mp4 not found")
    return FileResponse(INSTRUCTION_VIDEO_PATH, media_type="video/mp4")


@app.post("/extract")
async def extract(file: UploadFile = File(...)):
    if not (file.filename or "").lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Загрузите PDF файл (.pdf).")

    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl не установлен (нужен для XLSX).")

    pdf_bytes = await file.read()

    try:
        rows, stats_ = parse_items(pdf_bytes)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Не удалось распарсить PDF: {e}")

    if not rows:
        raise HTTPException(status_code=422, detail=f"Не удалось найти позиции. debug={stats_}")

    xlsx_bytes = make_xlsx(rows)
    increment_counter()
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="bau_items.xlsx"'},
    )


@app.post("/extract_async")
async def extract_async(file: UploadFile = File(...)):
    if not (file.filename or "").lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Загрузите PDF файл (.pdf).")

    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl не установлен (нужен для XLSX).")

    pdf_bytes = await file.read()
    original_filename = file.filename or "items.pdf"

    job_id = str(uuid4())
    _set_job(
        job_id,
        created_at=time.time(),
        status="processing",
        message="Старт обработки",
        processed_pages=0,
        total_pages=0,
        filename=(os.path.splitext(original_filename)[0] or "items") + ".xlsx",
    )

    def worker():
        try:
            _set_job(job_id, message="Читаю PDF…")
            rows, st = parse_items(pdf_bytes)

            _set_job(
                job_id,
                processed_pages=int(st.get("processed_pages", 0) or 0),
                total_pages=int(st.get("total_pages", 0) or 0),
                message="Формирую XLSX…",
            )

            if not rows:
                _set_job(job_id, status="error", message="Не удалось найти позиции", stats=st)
                return

            xlsx_bytes = make_xlsx(rows)
            _set_job_result(job_id, xlsx_bytes)
            increment_counter()

            _set_job(job_id, status="done", message="Готово", stats=st)
        except Exception as e:
            _set_job(job_id, status="error", message=f"Ошибка: {e}")

    threading.Thread(target=worker, daemon=True).start()
    return {"job_id": job_id}


@app.get("/job/{job_id}")
def job_status(job_id: str):
    j = _get_job(job_id)
    if not j:
        raise HTTPException(status_code=404, detail="job not found")

    return {
        "status": j.get("status"),
        "message": j.get("message"),
        "processed_pages": int(j.get("processed_pages", 0) or 0),
        "total_pages": int(j.get("total_pages", 0) or 0),
        "stats": j.get("stats"),
        "filename": j.get("filename"),
    }


import urllib.parse


@app.get("/job/{job_id}/download")
def job_download(job_id: str):
    j = _get_job(job_id)
    if not j:
        raise HTTPException(status_code=404, detail="job not found")
    if j.get("status") != "done":
        raise HTTPException(status_code=409, detail="job not done")

    xlsx_bytes = _get_job_result(job_id)
    if not xlsx_bytes:
        raise HTTPException(status_code=404, detail="result not found")

    filename_utf8 = j.get("filename", "items.xlsx")
    filename_ascii = re.sub(r"[^A-Za-z0-9_.-]+", "_", filename_utf8)
    quoted = urllib.parse.quote(filename_utf8)

    headers = {
        "Content-Disposition": (
            f'attachment; filename="{filename_ascii}"; '
            f"filename*=UTF-8''{quoted}"
        )
    }

    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
